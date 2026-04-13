import uuid
import tempfile
import threading
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file
import plotly.graph_objects as go
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from agent import RepositoryAnalyzerAgent

app = Flask(__name__)
api_key = 'sk-50D2tduetidrZ6GUqt0uzM5L72J5w6fKYeKsfAU4RADQK9pY'
agent = RepositoryAnalyzerAgent(api_key=api_key)

# Хранилище результатов (в реальном проекте следовало бы сохранять все в БД)
results = {}

def analyze_repository(repo_url: str) -> dict:
    """Анализ репозитория с помощью ИИ-агента."""
    json_llm = agent.analyze(repo_url)
    result = {
        "name": json_llm['name'],
        "url": repo_url,
        'last_commit': json_llm['last_commit'],
        'language': json_llm['language'],
        "analysis_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "model": "openrouter free",
        "criteria": json_llm["criteria"],
        "summary": json_llm["summary"],
        "charts": {
            "debt_by_type": json_llm["debt_by_type"],
            "debt_by_file": json_llm["debt_by_file"]
        }
    }
    return result


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/analyze', methods=['POST'])
def analyze():
    data = request.get_json()
    repo_url = data.get('repo_url')
    if not repo_url:
        return jsonify({'error': 'URL is required'}), 400

    task_id = str(uuid.uuid4())
    results[task_id] = {'status': 'pending', 'result': None}

    # Запуск анализа в фоновом потоке
    def run_analysis():
        try:
            result = analyze_repository(repo_url)
            results[task_id] = {'status': 'completed', 'result': result}
        except Exception as e:
            results[task_id] = {'status': 'error', 'error': str(e)}

    thread = threading.Thread(target=run_analysis)
    thread.start()

    return jsonify({'task_id': task_id})


@app.route('/status/<task_id>')
def status(task_id):
    task = results.get(task_id)
    if not task:
        return jsonify({'status': 'not_found'}), 404
    return jsonify(task)


@app.route('/report/<task_id>')
def view_report(task_id):
    task = results.get(task_id)
    if not task or task['status'] != 'completed':
        return "Отчёт не найден или ещё не готов", 404

    res = task['result']

    # Извлекаем данные для графиков
    debt_by_type = res['charts']['debt_by_type']
    debt_by_file = res['charts']['debt_by_file']

    # Столбчатая диаграмма
    fig1 = go.Figure(data=[
        go.Bar(
            x=debt_by_type['labels'],
            y=debt_by_type['values'],
            marker_color='rgb(55, 83, 109)'
        )
    ])
    fig1.update_layout(
        title='Количество фактов технического долга по типам',
        xaxis_title='Тип техдолга',
        yaxis_title='Количество',
        height=400,
        xaxis={'tickangle': -45, 'automargin': True},
        margin={'b': 100}
    )
    graph1_json = fig1.to_json()

    # Круговая диаграмма
    fig2 = go.Figure(data=[
        go.Pie(
            labels=debt_by_file['labels'],
            values=debt_by_file['values'],
            hole=0.3
        )
    ])
    fig2.update_layout(
        title='Распределение техдолга по файлам',
        height=400
    )
    graph2_json = fig2.to_json()

    return render_template('report.html',
                           result=res,
                           graph1_json=graph1_json,
                           graph2_json=graph2_json,
                           task_id=task_id)


@app.route('/download/<task_id>/<format>')
def download_report(task_id, format):
    task = results.get(task_id)
    if not task or task['status'] != 'completed':
        return "Отчёт не найден или ещё не готов", 404

    res = task['result']
    base_name = f"report_{res['name']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

    if format == 'excel':
        wb = openpyxl.Workbook()
        # Лист с общей информацией
        ws_info = wb.active
        ws_info.title = "Общая информация"
        info_data = [
            ["Параметр", "Значение"],
            ["Репозиторий", res['name']],
            ["URL", res['url']],
            ["Последний коммит", res['last_commit']],
            ["Язык", res['language']],
            ["Дата анализа", res['analysis_date']],
            ["Модель", res['model']],
            ["Итоговая оценка", res['summary']]
        ]
        for row in info_data:
            ws_info.append(row)

        # Лист с критериями
        ws_crit = wb.create_sheet("Критерии")
        ws_crit.append(["Критерий", "Оценка", "Комментарий"])
        for crit in res['criteria']:
            ws_crit.append([crit['name'], crit['score'], crit['comment']])

        # Лист с данными графиков
        ws_chart1 = wb.create_sheet("Техдолг по типам")
        df_type = pd.DataFrame(res['charts']['debt_by_type'])
        for r in dataframe_to_rows(df_type, index=False, header=True):
            ws_chart1.append(r)

        ws_chart2 = wb.create_sheet("Техдолг по файлам")
        df_file = pd.DataFrame(res['charts']['debt_by_file'])
        for r in dataframe_to_rows(df_file, index=False, header=True):
            ws_chart2.append(r)

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_excel:
            wb.save(tmp_excel.name)
            tmp_excel.flush()
            return send_file(
                tmp_excel.name,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f"{base_name}.xlsx"
            )
    else:
        return "Формат не поддерживается", 400


if __name__ == '__main__':
    app.run(debug=True)
